#!/usr/bin/env python3
import os
import re
import pandas as pd
from collections import defaultdict

# Try to import openpyxl and guide the user if it's not installed.
try:
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    print("The 'openpyxl' library is required to write formatted Excel files.")
    print("Please install it on your server by running: pip install openpyxl")
    exit()


def parse_champsim_file(filepath):
    """
    Parses a single ChampSim output file to extract specific metrics,
    focusing on IPC, LLC, and L2C statistics.

    Args:
        filepath (str): The full path to the ChampSim output file.

    Returns:
        dict: A dictionary containing the extracted metrics.
    """
    # Initialize a dictionary for the required metrics
    metrics = {
        "Trace File": os.path.basename(filepath),
        "IPC": None,
        "LLC Total Access": None,
        "LLC Total Hits": None,
        "LLC Total Misses": None,
        "L2C Total MPKI": None,
        "L2C Load Access": None,
        "L2C Load Hit": None,
        "L2C Load Miss": None,
        "L2C Load MPKI": None,
        "L2C Data Load MPKI": None,
        "L2C Prefetch Requested": None,
        "L2C Prefetch Issued": None,
        "L2C Prefetch Useful": None,
        "L2C Prefetch Useless": None,
        "L2C Useful Load Prefetches": None,
        "L2C Timely Prefetches": None,
        "L2C Late Prefetches": None,
        "L2C Dropped Prefetches": None,
        "L2C Average Miss Latency": None,
        "L2C Accuracy": None,
    }

    try:
        with open(filepath, 'r', errors='ignore') as f:
            content = f.read()

            # --- General Stats ---
            ipc_match = re.search(r"CPU 0 cumulative IPC:\s+([\d.]+)", content)
            if ipc_match:
                metrics["IPC"] = float(ipc_match.group(1))

            # --- LLC Stats ---
            llc_total_match = re.search(r"LLC TOTAL\s+ACCESS:\s+(\d+)\s+HIT:\s+(\d+)\s+MISS:\s+(\d+)", content)
            if llc_total_match:
                metrics["LLC Total Access"] = int(llc_total_match.group(1))
                metrics["LLC Total Hits"] = int(llc_total_match.group(2))
                metrics["LLC Total Misses"] = int(llc_total_match.group(3))

            # --- L2C Stats ---
            l2c_total_match = re.search(r"L2C TOTAL.*?MPKI:\s+([\d.]+)", content)
            if l2c_total_match:
                metrics["L2C Total MPKI"] = float(l2c_total_match.group(1))

            l2c_load_match = re.search(r"L2C LOAD\s+ACCESS:\s+(\d+)\s+HIT:\s+(\d+)\s+MISS:\s+(\d+).*?MPKI:\s+([\d.]+)", content)
            if l2c_load_match:
                metrics["L2C Load Access"] = int(l2c_load_match.group(1))
                metrics["L2C Load Hit"] = int(l2c_load_match.group(2))
                metrics["L2C Load Miss"] = int(l2c_load_match.group(3))
                metrics["L2C Load MPKI"] = float(l2c_load_match.group(4))

            l2c_data_load_mpki_match = re.search(r"L2C DATA LOAD MPKI:\s+([\d.]+)", content)
            if l2c_data_load_mpki_match:
                metrics["L2C Data Load MPKI"] = float(l2c_data_load_mpki_match.group(1))
            
            l2c_avg_miss_latency_match = re.search(r"L2C AVERAGE MISS LATENCY:\s+([\d.]+)", content)
            if l2c_avg_miss_latency_match:
                metrics["L2C Average Miss Latency"] = float(l2c_avg_miss_latency_match.group(1))

            # --- L2C Prefetch Stats ---
            l2c_prefetch_req_match = re.search(r"L2C PREFETCH\s+REQUESTED:\s+(\d+)\s+ISSUED:\s+(\d+)\s+USEFUL:\s+(\d+)\s+USELESS:\s+(\d+)", content)
            if l2c_prefetch_req_match:
                metrics["L2C Prefetch Requested"] = int(l2c_prefetch_req_match.group(1))
                metrics["L2C Prefetch Issued"] = int(l2c_prefetch_req_match.group(2))
                metrics["L2C Prefetch Useful"] = int(l2c_prefetch_req_match.group(3))
                metrics["L2C Prefetch Useless"] = int(l2c_prefetch_req_match.group(4))

            l2c_useful_load_match = re.search(r"L2C USEFUL LOAD PREFETCHES:\s+(\d+)", content)
            if l2c_useful_load_match:
                 metrics["L2C Useful Load Prefetches"] = int(l2c_useful_load_match.group(1))

            l2c_timely_match = re.search(r"L2C TIMELY PREFETCHES:\s+(\d+)\s+LATE PREFETCHES:\s+(\d+)\s+DROPPED PREFETCHES:\s+(\d+)", content)
            if l2c_timely_match:
                metrics["L2C Timely Prefetches"] = int(l2c_timely_match.group(1))
                metrics["L2C Late Prefetches"] = int(l2c_timely_match.group(2))
                metrics["L2C Dropped Prefetches"] = int(l2c_timely_match.group(3))

            l2c_accuracy_match = re.search(r"L2C .*? ACCURACY:\s+([\d.inf-]+)", content)
            if l2c_accuracy_match:
                accuracy_str = l2c_accuracy_match.group(1)
                try:
                    metrics["L2C Accuracy"] = float(accuracy_str)
                except ValueError:
                    metrics["L2C Accuracy"] = accuracy_str # Keep as string if 'inf' or '-nan'
    
    except IOError as e:
        print(f"Error reading file {filepath}: {e}")
        return None
        
    return metrics

def main():
    """
    Main function to find ChampSim files, parse them, and save them to a single
    Excel file with multiple sheets.
    """
    # --- CONFIGURATION ---
    # The script is in 'DATA/', so we go up one level to find 'result/'.
    RESULTS_DIR = "../result/"
    
    # Save the output file in the current directory ('DATA/')
    OUTPUT_FILE = "/home/neeraj/OneDrive/Assignment/collected_data.xlsx"
    # -------------------

    if not os.path.isdir(RESULTS_DIR):
        print(f"Error: Directory '{RESULTS_DIR}' not found.")
        print("Please ensure your 'result' directory is located one level above the 'DATA' directory.")
        return

    # This dictionary will hold dataframes, with sheet names as keys
    sheets_data = defaultdict(list)
    
    print(f"Starting scan in directory: '{RESULTS_DIR}'...")
    # Walk through the directory tree to collect all data
    for root, dirs, files in os.walk(RESULTS_DIR):
        # We only care about directories that contain .txt files
        txt_files = [f for f in files if f.endswith('.txt')]
        if not txt_files:
            continue

        # Create a sheet name from the relative path
        relative_path = os.path.relpath(root, RESULTS_DIR)
        # Handle case where files are directly in the results dir (e.g. '.')
        if relative_path == ".":
            continue
        sheet_name = relative_path.replace(os.sep, '_')
        
        print(f"Processing directory for sheet: {sheet_name}")

        for filename in sorted(txt_files):
            filepath = os.path.join(root, filename)
            metrics = parse_champsim_file(filepath)
            if metrics:
                sheets_data[sheet_name].append(metrics)

    if not sheets_data:
        print("No data was extracted. No file will be created.")
        return

    print(f"\nProcessing collected data and writing to {OUTPUT_FILE}...")
    
    try:
        with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
            for sheet_name, data_list in sorted(sheets_data.items()):
                print(f"Writing sheet: {sheet_name}")
                df = pd.DataFrame(data_list)
                df.to_excel(writer, sheet_name=sheet_name, index=False)

                # --- Formatting ---
                worksheet = writer.sheets[sheet_name]
                center_alignment = Alignment(horizontal='center', vertical='center')
                header_font = Font(bold=True, color="FFFFFF", size=12)
                data_font_bold = Font(bold=True)
                header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid") # Gray

                # Style all cells with center alignment
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.alignment = center_alignment
                
                # Style header row (Row 1)
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill

                # Style first column (Column A) to be bold
                for cell in worksheet['A']:
                    cell.font = data_font_bold

                # --- Auto-fit columns ---
                for col_idx in range(1, worksheet.max_column + 1):
                    column_letter = get_column_letter(col_idx)
                    max_length = 0
                    for cell in worksheet[column_letter]:
                        try:
                            if cell.value:
                                max_length = max(len(str(cell.value)), max_length)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        print(f"\nSuccessfully created Excel file: {OUTPUT_FILE}")

    except Exception as e:
        print(f"\nAn error occurred while writing the Excel file: {e}")

if __name__ == "__main__":
    main()

